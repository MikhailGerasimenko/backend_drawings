"""Выгрузка технологии изготовления в Excel (маршрутная + технологическая карта).

Шаблон — производственная форма (как в образце 07-54-319.xlsx): листы МК и ТК
в альбомной ориентации A4, с шапкой, маршрутом операций и переходами.
"""
from __future__ import annotations

import base64
import io
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from app.services.passport_normalize import passport_field_text
from app.services.technology_normalize import normalize_technology

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "assets"
    / "templates"
    / "technology_card.xlsx"
)

# Слоты операций на маршрутной карте (№ в колонке B, имя в D).
_MK1_SLOTS = (17, 21, 24, 27, 30, 33, 36, 39)  # лист 001, по 3 строки
_MK2_SLOTS = (8, 11, 14, 17, 20, 23, 26, 29, 32)  # лист 002 до блока «Мастер»
_MK2_MASTER_ROW = 37  # «Мастер / Подпись / ФИО / Дата» на продолжении МК
_TK1_BODY = (9, 34)  # строки содержания на первом листе ТК
_TK_CONT_BODY = (5, 36)  # строки содержания на листах-продолжениях ТК

# Диапазон твёрдости: «48...52», «48–52», «56,5»
_HARDNESS_RANGE = (
    r"\d+(?:[.,]\d+)?(?:\s*(?:\.\.\.|…|-|–|—)\s*\d+(?:[.,]\d+)?)?"
)
# И «HRC 48...52», и «48...52 HRC» (как в UI). Точки без цифр не считаем.
_HRC_RE = re.compile(
    rf"(?:HRC\s*(?:{_HARDNESS_RANGE})|(?:{_HARDNESS_RANGE})\s*HRC)\b",
    re.IGNORECASE,
)
_NBSP = "\xa0"


def _s(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _norm_diameter(text: str) -> str:
    """Привести обозначения диаметра к виду шаблона (‡).

    «Ф80» — диаметр; «Финальный» / «Фаска» — обычные слова, их не трогаем.
    """
    if not text:
        return text
    for ch in ("Ø", "⌀", "ø", "φ"):
        text = text.replace(ch, "‡")
    return re.sub(r"(?<![А-Яа-яA-Za-z])[Фф](?=\s*\d)", "‡", text)


def _material_without_hrc(material: str) -> str:
    """В шапке МК материал без блока HRC (твёрдость в отдельной ячейке)."""
    if not material:
        return ""
    cleaned = _HRC_RE.sub("", material)
    cleaned = re.sub(r"[,;]\s*$", "", cleaned.strip())
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" ,;")
    return cleaned


def _extract_hardness(*parts: str) -> str:
    """Вернуть твёрдость в виде «48...52 HRC» (как в ключевых размерах UI)."""
    for part in parts:
        if not part:
            continue
        m = _HRC_RE.search(part)
        if not m:
            continue
        raw = m.group(0).upper().replace("…", "...").replace("–", "...").replace("—", "...")
        raw = re.sub(r"\s*-\s*", "...", raw)
        raw = re.sub(r"\s+", " ", raw).strip()
        nums = re.sub(r"\s*HRC\s*", " ", raw, flags=re.IGNORECASE).strip()
        if not nums:
            continue
        return f"{nums} HRC"
    return ""


def _split_name(name: str, max_first: int = 22) -> tuple[str, str]:
    """Разбить длинное имя операции на 1–2 строки как в образце."""
    name = _s(name)
    if not name:
        return "", ""
    if len(name) <= max_first:
        return name, ""
    # Предпочитаем разрыв по пробелу/дефису
    cut = name.rfind(" ", 0, max_first + 1)
    if cut < 8:
        cut = name.rfind("-", 0, max_first + 1)
    if cut < 8:
        cut = max_first
    return name[:cut].rstrip("- "), name[cut:].lstrip("- ")


def _wrap_line(text: str, width: int = 95) -> list[str]:
    text = _s(text)
    if not text:
        return []
    out: list[str] = []
    for para in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        para = para.strip()
        if not para:
            continue
        while len(para) > width:
            cut = para.rfind(" ", 0, width + 1)
            if cut < width // 3:
                cut = width
            out.append(para[:cut].rstrip())
            para = para[cut:].lstrip()
        if para:
            out.append(para)
    return out


def _set(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    horizontal: str | None = None,
    bold: bool | None = None,
) -> None:
    cell = ws.cell(row, col)
    # openpyxl: писать можно только в верхний левый угол merge
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                cell = ws.cell(merged.min_row, merged.min_col)
                break
    cell.value = value if value not in ("", None) else None
    if horizontal:
        prev = cell.alignment
        cell.alignment = Alignment(
            horizontal=horizontal,
            vertical=prev.vertical or "center",
            wrap_text=True,
            textRotation=prev.textRotation,
            shrinkToFit=False,
            indent=prev.indent,
        )
    if bold is not None:
        font = copy(cell.font)
        font.bold = bold
        cell.font = font


def _enable_wrap(ws: Worksheet) -> None:
    """Перенос текста во всех ячейках листа (чтобы длинные значения не вылезали)."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            prev = cell.alignment
            cell.alignment = Alignment(
                horizontal=prev.horizontal,
                vertical=prev.vertical or "center",
                wrap_text=True,
                textRotation=prev.textRotation,
                shrinkToFit=False,
                indent=prev.indent,
            )


def _clear_mk_slots(ws: Worksheet, slots: tuple[int, ...], span: int = 3) -> None:
    for start in slots:
        for r in range(start, start + span):
            for col in (2, 4, 11, 15):
                _set(ws, r, col, None)


def _clear_tk_body(ws: Worksheet, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for col in (2, 3, 5, 14):  # B, C, E, N
            _set(ws, r, col, None)

def _route_steps(tj: dict) -> list[dict]:
    route = tj.get("route") or []
    steps: list[dict] = []
    for i, step in enumerate(route):
        if not isinstance(step, dict):
            continue
        num = step.get("number")
        try:
            num_i = int(num) if num is not None else i + 1
        except (TypeError, ValueError):
            num_i = i + 1
        name = _s(step.get("name")) or f"Операция {num_i}"
        steps.append(
            {
                "number": num_i,
                "name": name,
                "equipment": _s(step.get("equipment")),
                "transitions": _s(step.get("transitions")),
                "final_sizes": _s(step.get("final_sizes")),
            }
        )
    return steps


def _fill_mk_header(
    ws: Worksheet,
    *,
    designation: str,
    part_name: str,
    material: str,
    blank_dims: str,
    blank_weight: str,
    finish_weight: str,
    hardness: str,
) -> None:
    _set(ws, 6, 4, part_name or None)  # D6 Деталь
    _set(ws, 4, 2, None)  # B4 Цех-заказчик — нет данных, не заполняем
    _set(ws, 6, 2, None)  # B6 значение цеха из шаблона (КлЦ)
    _set(ws, 9, 6, _norm_diameter(blank_dims) or None)  # F9 размер заготовки
    _set(ws, 9, 10, blank_weight or None)  # J9 вес заготовки
    _set(ws, 9, 12, finish_weight or None)  # L9 чистовой вес
    _set(ws, 9, 15, hardness or None)  # O9 твердость
    _set(ws, 10, 2, designation or None)  # B10 № чертежа
    _set(ws, 10, 4, material or None)  # D10 материал


def _fill_mk_ops(
    ws: Worksheet, slots: tuple[int, ...], steps: list[dict], nh_col: int
) -> None:
    for start, step in zip(slots, steps):
        _set(ws, start, 2, step["number"], horizontal="center")
        line1, line2 = _split_name(step["name"])
        _set(ws, start, 4, line1)
        if line2:
            _set(ws, start + 1, 4, line2)
        _set(ws, start, nh_col, _NBSP)


def _fill_tk_header(
    ws: Worksheet,
    *,
    designation: str,
    part_name: str,
    blank_type: str,
    material: str,
    blank_dims: str,
    qty: str,
    blank_weight: str,
    piece_weight: str,
    sheet_no: int,
    sheets_total: int,
) -> None:
    _set(ws, 2, 9, designation or None)  # I2
    _set(ws, 3, 9, part_name or None)  # I3
    _set(ws, 4, 15, sheet_no)  # O4 Лист
    _set(ws, 5, 15, sheets_total)  # O5 Листов
    _set(ws, 6, 2, blank_type or None)
    _set(ws, 6, 5, material or None)
    _set(ws, 6, 6, _norm_diameter(blank_dims) or None)
    _set(ws, 6, 11, qty or None)
    _set(ws, 6, 12, blank_weight or None)
    _set(ws, 6, 13, piece_weight or None)


def _fill_tk_cont_header(
    ws: Worksheet,
    *,
    designation: str,
    part_name: str,
    sheet_no: int,
    sheets_total: int,
) -> None:
    _set(ws, 2, 4, designation or None)  # D2
    _set(ws, 2, 6, sheet_no)  # F2 Лист
    _set(ws, 3, 4, part_name or None)  # D3
    _set(ws, 3, 6, sheets_total)  # F3 Листов


def _tk_lines_for_steps(
    steps: list[dict],
) -> list[tuple[int | None, str, str | None]]:
    """Строки ТК: (номер|None, текст, оборудование|None)."""
    lines: list[tuple[int | None, str, str | None]] = []
    for step in steps:
        lines.append((step["number"], step["name"], step["equipment"] or None))
        body_parts: list[str] = []
        if step["transitions"]:
            body_parts.extend(_wrap_line(step["transitions"], 100))
        if step["final_sizes"]:
            body_parts.extend(
                _wrap_line(f"Итоговые размеры: {step['final_sizes']}", 100)
            )
        for t in body_parts:
            lines.append((None, _norm_diameter(t), None))
    return lines


def _write_tk_page(
    ws: Worksheet,
    lines: list[tuple[int | None, str, str | None]],
    *,
    start_row: int,
    end_row: int,
    num_col: int,
    text_col: int,
    eq_col: int,
) -> list[tuple[int | None, str, str | None]]:
    """Записать строки в диапазон; вернуть остаток."""
    row = start_row
    idx = 0
    while idx < len(lines) and row <= end_row:
        num, text, eq = lines[idx]
        if num is not None:
            _set(ws, row, num_col, num, horizontal="center", bold=False)
            _set(ws, row, text_col, text, horizontal="left", bold=True)
        else:
            _set(ws, row, text_col, text, horizontal="left", bold=False)
        if eq:
            _set(ws, row, eq_col, eq, horizontal="center", bold=False)
        idx += 1
        row += 1
    return lines[idx:]


def _tk_body_capacity(start_row: int, end_row: int) -> int:
    return end_row - start_row + 1


def _tk_page_count(n_lines: int) -> int:
    first = _tk_body_capacity(*_TK1_BODY)
    cont = _tk_body_capacity(*_TK_CONT_BODY)
    if n_lines <= first:
        return 1
    extra = n_lines - first
    return 1 + (extra + cont - 1) // cont


def _continuation_name(wb, prefix: str) -> str:
    i = 2
    while f"{prefix}-{i}" in wb.sheetnames:
        i += 1
    return f"{prefix}-{i}"


def _place_after(wb, ws: Worksheet, after_name: str) -> None:
    if after_name not in wb.sheetnames:
        return
    target = wb.sheetnames.index(after_name) + 1
    current = wb.sheetnames.index(ws.title)
    if current != target:
        wb.move_sheet(ws, offset=target - current)


def _clear_mk_cont_slots(ws: Worksheet) -> None:
    for start in _MK2_SLOTS:
        for r in range(start, start + 3):
            for col in (2, 4, 11):
                _set(ws, r, col, None)


def _origin_cell(ws: Worksheet, row: int, col: int):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(merged.min_row, merged.min_col)
    return cell


def _copy_cell_look(src, dst) -> None:
    dst.font = copy(src.font)
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)


def _unmerge_block(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    overlapping = [
        str(rng)
        for rng in list(ws.merged_cells.ranges)
        if not (
            rng.max_row < min_row
            or rng.min_row > max_row
            or rng.max_col < min_col
            or rng.min_col > max_col
        )
    ]
    for ref in overlapping:
        ws.unmerge_cells(ref)


def _master_src_col(dest_col: int, last_col: int) -> int:
    """Колонки B–H как на 002; хвост I…last как пустая ячейка I–L."""
    if dest_col <= 8:
        return dest_col
    if dest_col == 9:
        return 9
    if dest_col == last_col:
        return 12
    return 10


def _apply_mk_master_footer(
    dest: Worksheet, src: Worksheet, *, dest_row: int, last_col: int
) -> None:
    """Перенести блок «Мастер / Подпись / ФИО / Дата» на последний лист МК."""
    src_row = _MK2_MASTER_ROW
    r1, r2 = dest_row, dest_row + 1
    _unmerge_block(dest, r1, r2, 2, last_col)
    dest.merge_cells(start_row=r1, start_column=2, end_row=r2, end_column=4)
    dest.merge_cells(start_row=r1, start_column=6, end_row=r1, end_column=7)
    dest.merge_cells(start_row=r2, start_column=6, end_row=r2, end_column=7)
    dest.merge_cells(start_row=r1, start_column=9, end_row=r2, end_column=last_col)
    for offset in (0, 1):
        for col in range(2, last_col + 1):
            src_cell = _origin_cell(src, src_row + offset, _master_src_col(col, last_col))
            dst_cell = dest.cell(r1 + offset, col)
            if isinstance(dst_cell, MergedCell):
                continue
            _copy_cell_look(src_cell, dst_cell)
    dest.row_dimensions[r1].height = src.row_dimensions[src_row].height or 17.1
    dest.row_dimensions[r2].height = src.row_dimensions[src_row + 1].height or 17.1
    _set(dest, r1, 2, "Мастер", horizontal="right")
    _set(dest, r1, 5, "Подпись", horizontal="center")
    _set(dest, r1, 6, "ФИО", horizontal="center")
    _set(dest, r1, 8, "Дата", horizontal="center")
    # подписи пустые — только рамки
    _set(dest, r2, 5, None)
    _set(dest, r2, 6, None)
    _set(dest, r2, 8, None)


def _clone_sheet_style(src: Worksheet, dst: Worksheet) -> None:
    """Копировать ширины колонок и высоты строк."""
    for letter, dim in src.column_dimensions.items():
        dst.column_dimensions[letter].width = dim.width
    for idx, dim in src.row_dimensions.items():
        if dim.height is not None:
            dst.row_dimensions[idx].height = dim.height
    dst.page_setup.orientation = src.page_setup.orientation
    dst.page_setup.paperSize = src.page_setup.paperSize
    dst.page_margins.left = src.page_margins.left
    dst.page_margins.right = src.page_margins.right
    dst.page_margins.top = src.page_margins.top
    dst.page_margins.bottom = src.page_margins.bottom
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight


def _append_mk_overflow_sheet(
    wb, template_ws: Worksheet, steps: list[dict], after_name: str
) -> str:
    """Доп. лист маршрутной карты, если операции не влезли на 001–002."""
    name = _continuation_name(wb, "002")
    ws = wb.copy_worksheet(template_ws)
    ws.title = name
    _clone_sheet_style(template_ws, ws)
    _clear_mk_cont_slots(ws)
    _fill_mk_ops(ws, _MK2_SLOTS, steps, nh_col=11)
    _place_after(wb, ws, after_name)
    return ws.title


def _append_tk_overflow_sheet(
    wb,
    template_ws: Worksheet,
    *,
    designation: str,
    part_name: str,
    sheet_no: int,
    sheets_total: int,
    lines: list[tuple[int | None, str, str | None]],
    after_name: str,
) -> tuple[list[tuple[int | None, str, str | None]], str]:
    """Создать дополнительный лист продолжения ТК по образцу листа 004."""
    wb_name = _continuation_name(wb, "004")
    ws = wb.copy_worksheet(template_ws)
    ws.title = wb_name
    _clone_sheet_style(template_ws, ws)
    _fill_tk_cont_header(
        ws,
        designation=designation,
        part_name=part_name,
        sheet_no=sheet_no,
        sheets_total=sheets_total,
    )
    _clear_tk_body(ws, *_TK_CONT_BODY)
    rest = _write_tk_page(
        ws,
        lines,
        start_row=_TK_CONT_BODY[0],
        end_row=_TK_CONT_BODY[1],
        num_col=2,
        text_col=3,
        eq_col=5,
    )
    _place_after(wb, ws, after_name)
    return rest, ws.title


def build_technology_xlsx(
    title: str,
    technology_json: dict | None,
    passport: dict | None = None,
    technology_text: str | None = None,
) -> str:
    """Собрать .xlsx и вернуть data:application/...;base64,..."""
    if technology_json:
        _, tj = normalize_technology(technology_json, passport)
    else:
        # fallback: одна операция из текста
        tj = {
            "schema_version": "2.0",
            "header": {
                "part_designation": title or "",
                "part_name": "",
                "material": "",
            },
            "blank": {},
            "route": [
                {
                    "number": 1,
                    "name": "Технология",
                    "transitions": technology_text or "Нет данных",
                    "equipment": "",
                    "final_sizes": "",
                }
            ],
            "metadata": {},
            "heat_treatment": "",
            "key_dimensions": "",
        }

    header = tj.get("header") or {}
    blank = tj.get("blank") or {}
    steps = _route_steps(tj)
    if not steps:
        steps = [
            {
                "number": 1,
                "name": "Технология",
                "equipment": "",
                "transitions": technology_text or "Нет данных",
                "final_sizes": "",
            }
        ]

    designation = _s(header.get("part_designation")) or _s(title)
    part_name = _s(header.get("part_name"))
    material = _s(header.get("material"))
    material_mk = _material_without_hrc(material) or material
    blank_type = _s(blank.get("type"))
    blank_dims = _s(blank.get("dimensions"))
    if blank.get("allowances") and blank_dims:
        # в образце размер и кол-во в одной ячейке: ‡80х36/1шт
        pass
    hardness = _extract_hardness(
        _s(tj.get("key_dimensions")),
        _s(tj.get("heat_treatment")),
        material,
        passport_field_text(passport, "material_hardness"),
    )
    finish_weight = passport_field_text(passport, "mass")
    blank_weight = ""
    piece_weight = finish_weight
    qty = "1"
    if blank_dims and "/" not in blank_dims:
        blank_dims_mk = f"{blank_dims}\n/1шт" if blank_dims else ""
    else:
        blank_dims_mk = blank_dims

    if not TEMPLATE_PATH.is_file():
        raise FileNotFoundError(f"Шаблон не найден: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws_mk1 = wb["001"]
    ws_mk2 = wb["002"]
    ws_tk1 = wb["003"]
    ws_tk2 = wb["004"]

    # --- Маршрутная карта: столько листов, сколько нужно ---
    _clear_mk_slots(ws_mk1, _MK1_SLOTS, span=3)
    _fill_mk_header(
        ws_mk1,
        designation=designation,
        part_name=part_name,
        material=material_mk,
        blank_dims=blank_dims_mk,
        blank_weight=blank_weight,
        finish_weight=finish_weight,
        hardness=hardness,
    )
    mk1_count = len(_MK1_SLOTS)
    _fill_mk_ops(ws_mk1, _MK1_SLOTS, steps[:mk1_count], nh_col=15)
    rest_mk = steps[mk1_count:]
    if not rest_mk:
        # Блок «Мастер» в шаблоне только на 002. Если маршрут влез в 001 —
        # переносим подпись вниз первого листа, иначе при удалении 002 она пропадает.
        if len(steps) < mk1_count:
            _apply_mk_master_footer(ws_mk1, ws_mk2, dest_row=39, last_col=16)
            wb.remove(ws_mk2)
        else:
            _clear_mk_cont_slots(ws_mk2)
    else:
        _clear_mk_cont_slots(ws_mk2)
        _fill_mk_ops(ws_mk2, _MK2_SLOTS, rest_mk[: len(_MK2_SLOTS)], nh_col=11)
        rest_mk = rest_mk[len(_MK2_SLOTS) :]
        last_mk = "002"
        while rest_mk:
            chunk = rest_mk[: len(_MK2_SLOTS)]
            last_mk = _append_mk_overflow_sheet(wb, ws_mk2, chunk, last_mk)
            rest_mk = rest_mk[len(_MK2_SLOTS) :]

    # --- Технологическая карта: только маршрут, без пустых листов шаблона ---
    tk_lines = _tk_lines_for_steps(steps)
    sheets_total = _tk_page_count(len(tk_lines))

    _fill_tk_header(
        ws_tk1,
        designation=designation,
        part_name=part_name,
        blank_type=blank_type,
        material=material_mk,
        blank_dims=blank_dims,
        qty=qty,
        blank_weight=blank_weight or piece_weight,
        piece_weight=piece_weight,
        sheet_no=1,
        sheets_total=sheets_total,
    )
    _clear_tk_body(ws_tk1, *_TK1_BODY)
    rest = _write_tk_page(
        ws_tk1,
        tk_lines,
        start_row=_TK1_BODY[0],
        end_row=_TK1_BODY[1],
        num_col=2,
        text_col=3,
        eq_col=14,
    )

    # Подпись: только подписи полей, без ФИО и даты
    _set(ws_tk1, 35, 2, "Технолог:")
    _set(ws_tk1, 35, 4, None)
    _set(ws_tk1, 35, 9, "Дата:")
    _set(ws_tk1, 35, 10, None)

    if not rest:
        wb.remove(ws_tk2)
    else:
        _fill_tk_cont_header(
            ws_tk2,
            designation=designation,
            part_name=part_name,
            sheet_no=2,
            sheets_total=sheets_total,
        )
        _clear_tk_body(ws_tk2, *_TK_CONT_BODY)
        rest = _write_tk_page(
            ws_tk2,
            rest,
            start_row=_TK_CONT_BODY[0],
            end_row=_TK_CONT_BODY[1],
            num_col=2,
            text_col=3,
            eq_col=5,
        )
        last_tk = "004"
        sheet_no = 3
        while rest:
            rest, last_tk = _append_tk_overflow_sheet(
                wb,
                ws_tk2,
                designation=designation,
                part_name=part_name,
                sheet_no=sheet_no,
                sheets_total=sheets_total,
                lines=rest,
                after_name=last_tk,
            )
            sheet_no += 1

    for ws in wb.worksheets:
        _enable_wrap(ws)

    buf = io.BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return (
        "data:application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet;base64," + b64
    )
