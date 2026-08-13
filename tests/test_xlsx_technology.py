"""Выгрузка технологии в Excel (маршрутная + технологическая карта)."""
import base64
import io

from openpyxl import load_workbook

from app.services.xlsx_technology import (
    _extract_hardness,
    _norm_diameter,
    _split_name,
    build_technology_xlsx,
)


def _decode_xlsx(data_url: str):
    assert data_url.startswith(
        "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
    )
    raw = base64.b64decode(data_url.split(",", 1)[1])
    return load_workbook(io.BytesIO(raw))


def test_helpers():
    assert _norm_diameter("Ø30 H10") == "‡30 H10"
    assert _extract_hardness("Сталь 9ХС, HRC 48...52") == "48...52 HRC"
    assert (
        _extract_hardness(
            "Габариты: Ø20,4(+0,2) х 14,9±0,1 мм\n"
            "Отверстие: Ø6,017(-0,003 / -0,026), Ra 0,8\n"
            "Твердость: 48...52 HRC\n"
            "Шероховатость торцов: Ra 3,2"
        )
        == "48...52 HRC"
    )
    assert _extract_hardness("HRC.") == ""
    assert _extract_hardness("HRC .") == ""
    assert _split_name("Ленточно (проволочно)-отрезная")[0]


def test_build_technology_xlsx_fills_template():
    tj = {
        "schema_version": "2.0",
        "header": {
            "part_designation": "07-54-319",
            "part_name": "Проставка",
            "material": "Сталь 9ХС HRC 48...52",
            "features": "",
        },
        "key_dimensions": "‡69,72 / ‡30H10",
        "blank": {
            "type": "Прокат",
            "dimensions": "Ø80х36",
            "allowances": "1,5 мм",
        },
        "route": [
            {
                "code": "OP01",
                "number": 1,
                "name": "Ленточно (проволочно)-отрезная",
                "equipment": "H-460HANC",
                "transitions": "Отрезать заготовку, выдерживая размер L=36+/-1 мм",
                "final_sizes": "L=36",
            },
            {
                "code": "OP02",
                "number": 2,
                "name": "Отжиг",
                "equipment": "СНЗ-6.12.4/12М1",
                "transitions": "",
                "final_sizes": "",
            },
            {
                "code": "OP03",
                "number": 3,
                "name": "Токарная",
                "equipment": "16К20",
                "transitions": "Установить, выверить и закрепить заготовку\nПодрезать торец как чисто",
                "final_sizes": "30(-0,05)",
            },
        ],
        "heat_treatment": "HRC 48...52",
        "finish_after_heat_treatment": "",
        "confirmation_required": "",
        "dimensions_control": "Контроль размеров по чертежу",
        "metadata": {
            "card_version": "draft v1.0",
            "author": "ИИ-ассистент",
            "date": "2026-08-11",
            "files_used": ["чертёж"],
            "allowance_rule_version": "v1.1",
        },
    }
    passport = {
        "schema_version": "2.0",
        "mass": {"value": "0,76", "missing_on_drawing": False},
        "material_hardness": {
            "value": "Сталь 9ХС, HRC 48...52",
            "missing_on_drawing": False,
        },
    }

    data_url = build_technology_xlsx("07-54-319 - Проставка", tj, passport=passport)
    wb = _decode_xlsx(data_url)

    assert wb.sheetnames == ["001", "003"]
    mk = wb["001"]
    assert mk["B10"].value == "07-54-319"
    assert mk["D6"].value == "Проставка"
    assert "9ХС" in str(mk["D10"].value)
    assert mk["O9"].value == "48...52 HRC"
    assert mk["B4"].value in (None, "")
    assert mk["B6"].value in (None, "")
    assert mk["F9"].alignment.wrap_text is True
    assert mk["B17"].value == 1
    assert "Ленточно" in str(mk["D17"].value)
    assert mk["L9"].value == "0,76"

    tk = wb["003"]
    assert tk["I2"].value == "07-54-319"
    assert tk["I3"].value == "Проставка"
    assert tk["B9"].value == 1
    assert tk["B9"].alignment.horizontal == "center"
    assert "Ленточно" in str(tk["C9"].value)
    assert tk["C9"].font.bold is True
    assert tk["C9"].alignment.horizontal == "left"
    assert tk["C10"].font.bold is False
    assert tk["C10"].alignment.horizontal == "left"
    assert "Итоговые размеры" in str(tk["C11"].value)
    assert tk["C11"].font.bold is False
    assert tk["N9"].value == "H-460HANC"
    assert tk["B35"].value == "Технолог:"
    assert tk["D35"].value in (None, "")
    assert tk["I35"].value == "Дата:"
    assert tk["J35"].value in (None, "")
    dumped = " ".join(
        str(c.value)
        for ws in wb.worksheets
        if ws.title != "001"
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row)
        for c in row
        if c.value
    )
    assert "Контроль размеров по чертежу" not in dumped
    assert "Термообработка:" not in dumped
    assert "ИИ-ассистент" not in dumped
    assert "draft v1.0" not in dumped


def test_build_technology_xlsx_overflow_extra_sheet():
    route = []
    for i in range(1, 16):
        route.append(
            {
                "code": f"OP{i:02d}",
                "number": i,
                "name": f"Операция длинная {i}",
                "equipment": f"EQ-{i}",
                "transitions": "Шаг A\nШаг B\nШаг C\nШаг D\nШаг E\nШаг F",
                "final_sizes": "size",
            }
        )
    tj = {
        "schema_version": "2.0",
        "header": {
            "part_designation": "X-1",
            "part_name": "Деталь",
            "material": "Сталь 45",
        },
        "blank": {"type": "Пруток", "dimensions": "Ø50", "allowances": ""},
        "route": route,
        "metadata": {"author": "test", "date": "2026-01-01"},
    }
    data_url = build_technology_xlsx("X-1", tj)
    wb = _decode_xlsx(data_url)
    assert "001" in wb.sheetnames
    assert "002" in wb.sheetnames  # 15 операций > 8 слотов первого листа МК
    assert "003" in wb.sheetnames
    assert "004" in wb.sheetnames  # длинные переходы не влезают на первый лист ТК
    assert len(wb.sheetnames) >= 4


def test_xlsx_pages_scale_with_route_size():
    def _card(n_ops: int, transitions: str = "Шаг"):
        return {
            "schema_version": "2.0",
            "header": {
                "part_designation": "X",
                "part_name": "Деталь",
                "material": "Сталь 45",
            },
            "blank": {"type": "Пруток", "dimensions": "Ø50"},
            "route": [
                {
                    "number": i,
                    "name": f"Операция {i}",
                    "equipment": "EQ",
                    "transitions": transitions,
                    "final_sizes": "",
                }
                for i in range(1, n_ops + 1)
            ],
        }

    small = _decode_xlsx(build_technology_xlsx("X", _card(3)))
    assert "002" not in small.sheetnames
    assert "004" not in small.sheetnames
    assert small.sheetnames == ["001", "003"]

    mk_two = _decode_xlsx(build_technology_xlsx("X", _card(9)))
    assert "002" in mk_two.sheetnames
    assert "001" in mk_two.sheetnames

    mk_extra = _decode_xlsx(build_technology_xlsx("X", _card(20)))
    assert "002" in mk_extra.sheetnames
    assert any(name.startswith("002-") for name in mk_extra.sheetnames)
